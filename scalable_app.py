
from flask import Flask, render_template, request, redirect, flash, send_file
import pandas as pd
from datetime import datetime
import os

app = Flask(__name__)
app.secret_key = 'your_secret_key'

EXCEL_FILE = 'Outpatient_Report.xlsx'
SUMMARY_FILE = 'Summary_Report.xlsx'
DATE_COL_NAME = 'DATE'

def get_sheet_name_from_date(date_obj):
    return date_obj.strftime('%b_%Y')

def load_or_create_sheet(sheet_name):
    if not os.path.exists(EXCEL_FILE):
        columns = [DATE_COL_NAME,
                   'Old_M_M', 'Old_F_M', 'Old_C_M', 'New_M_M', 'New_F_M', 'New_C_M', 'Total_Morning',
                   'Old_M_E', 'Old_F_E', 'Old_C_E', 'New_M_E', 'New_F_E', 'New_C_E', 'Total_Evening',
                   'Grand_Total']
        df = pd.DataFrame(columns=columns)
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    try:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet_name)
    except:
        columns = [DATE_COL_NAME,
                   'Old_M_M', 'Old_F_M', 'Old_C_M', 'New_M_M', 'New_F_M', 'New_C_M', 'Total_Morning',
                   'Old_M_E', 'Old_F_E', 'Old_C_E', 'New_M_E', 'New_F_E', 'New_C_E', 'Total_Evening',
                   'Grand_Total']
        df = pd.DataFrame(columns=columns)
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)
    return df

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        date_str = request.form.get('date')
        try:
            date_obj = pd.to_datetime(date_str).normalize()
        except:
            flash("Invalid date format.")
            return redirect('/')

        def parse_int(value):
            try:
                return int(value)
            except:
                return 0

        # Morning session values
        old_m_m = parse_int(request.form.get('old_m_m'))
        old_f_m = parse_int(request.form.get('old_f_m'))
        old_c_m = parse_int(request.form.get('old_c_m'))
        new_m_m = parse_int(request.form.get('new_m_m'))
        new_f_m = parse_int(request.form.get('new_f_m'))
        new_c_m = parse_int(request.form.get('new_c_m'))
        total_m = old_m_m + old_f_m + old_c_m + new_m_m + new_f_m + new_c_m

        # Evening session values
        old_m_e = parse_int(request.form.get('old_m_e'))
        old_f_e = parse_int(request.form.get('old_f_e'))
        old_c_e = parse_int(request.form.get('old_c_e'))
        new_m_e = parse_int(request.form.get('new_m_e'))
        new_f_e = parse_int(request.form.get('new_f_e'))
        new_c_e = parse_int(request.form.get('new_c_e'))
        total_e = old_m_e + old_f_e + old_c_e + new_m_e + new_f_e + new_c_e

        grand_total = total_m + total_e

        sheet_name = get_sheet_name_from_date(date_obj)
        df = load_or_create_sheet(sheet_name)

        df[DATE_COL_NAME] = pd.to_datetime(df[DATE_COL_NAME], errors='coerce').dt.normalize()
        row_index = df[df[DATE_COL_NAME] == date_obj].index

        if not row_index.empty:
            idx = row_index[0]
            df.loc[idx, [
                'Old_M_M', 'Old_F_M', 'Old_C_M', 'New_M_M', 'New_F_M', 'New_C_M',
                'Old_M_E', 'Old_F_E', 'Old_C_E', 'New_M_E', 'New_F_E', 'New_C_E',
                'Total_Morning', 'Total_Evening', 'Grand_Total'
            ]] = [
                old_m_m, old_f_m, old_c_m, new_m_m, new_f_m, new_c_m,
                old_m_e, old_f_e, old_c_e, new_m_e, new_f_e, new_c_e,
                total_m, total_e, grand_total
            ]
            flash("Data updated successfully.")
        else:
            new_row = {
                DATE_COL_NAME: date_obj,
                'Old_M_M': old_m_m, 'Old_F_M': old_f_m, 'Old_C_M': old_c_m,
                'New_M_M': new_m_m, 'New_F_M': new_f_m, 'New_C_M': new_c_m,
                'Total_Morning': total_m,
                'Old_M_E': old_m_e, 'Old_F_E': old_f_e, 'Old_C_E': old_c_e,
                'New_M_E': new_m_e, 'New_F_E': new_f_e, 'New_C_E': new_c_e,
                'Total_Evening': total_e,
                'Grand_Total': grand_total
            }
            df = pd.concat([df, pd.DataFrame([new_row])], ignore_index=True)
            flash("New row added successfully.")

        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        return redirect('/')

    return render_template('form_cases_sessions.html')

@app.route('/download_summary')
def download_summary():
    from openpyxl import load_workbook

    if not os.path.exists(EXCEL_FILE):
        flash("No data available yet.")
        return redirect('/')

    wb = load_workbook(EXCEL_FILE)
    summary_data = []

    for sheet in wb.sheetnames:
        df = pd.read_excel(EXCEL_FILE, sheet_name=sheet)
        for col in ['Total_Morning', 'Total_Evening', 'Grand_Total']:
            df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
        totals = {
            "Month": sheet,
            "Morning_Total": df['Total_Morning'].sum(),
            "Evening_Total": df['Total_Evening'].sum(),
            "Grand_Total": df['Grand_Total'].sum()
        }
        summary_data.append(totals)

    summary_df = pd.DataFrame(summary_data)
    summary_df.to_excel(SUMMARY_FILE, index=False)
    return send_file(SUMMARY_FILE, as_attachment=True)

if __name__ == '__main__':
    app.run(debug=True)
